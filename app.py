from flask import Flask, jsonify,request
from flask_sqlalchemy import SQLAlchemy
from flask_marshmallow import Marshmallow
from openpyxl import load_workbook

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///data.db'

db = SQLAlchemy(app)
ma = Marshmallow(app)

class Post(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100))
    age = db.Column(db.Float)
    country = db.Column(db.String(50))
 
    def __init__(self,name, age,country):
        # self.id = id
        self.name = name
        self.age = age
        self.country = country
        
class PostSchema(ma.Schema):
    class Meta:
        fields = ("id", "name", "age","country")
        
post_schema = PostSchema()
posts_schema = PostSchema(many=True)
    
@app.route('/get', methods= ['POST'])
def add_data():
    if request.method == 'POST':
        data = request.files['afroj']
        afroj =load_workbook(data)
        afroj1=afroj.active
        for i in afroj1.iter_rows(min_row=2,values_only = True):
            data1 = Post(name=i[0],age=i[1],country=i[2])
            db.session.add(data1)
        db.session.commit()
    return "msg:data retrieve"

@app.route('/retrieve', methods = ['GET'])
def get_post():
    all_posts = Post.query.all()
    result = posts_schema.dumps(all_posts)
    
    return jsonify(result)

@app.route('/get_details/<int:id>', methods = ['GET'])
def get_details(id):
   
    post = Post.query.filter_by(id=id).first()
    result = post_schema.dumps(post)
    
    return jsonify(result)
    

@app.route('/post_delete/<int:id>/', methods = ['DELETE'])
def post_delete(id):
    post = Post.query.get(id)
    db.session.delete(post)
    db.session.commit()
    print(post)
    return post_schema.jsonify(post)

@app.route('/post_updates/<int:id>/', methods = ['PUT'])
def post_update(id):
    post = Post.query.get(id)
  
    name = request.json['name']
    age = request.json['age']
    country = request.json['country']
    
    
    post.name = name
    post.age = age
    post.country = country
    db.session.commit()
    return post_schema.jsonify(post)

with app.app_context():
    db.create_all()
    
if __name__ == "__main__":
     app.run(debug=True ,port=8000,use_reloader=False)
    